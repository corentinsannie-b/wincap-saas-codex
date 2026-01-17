"""Excel styling constants for the Databook."""

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

# Colors (Wincap-style blue theme)
DARK_BLUE = "1F4E79"
MEDIUM_BLUE = "2E75B6"
LIGHT_BLUE = "BDD7EE"
WHITE = "FFFFFF"
LIGHT_GRAY = "F2F2F2"
BLACK = "000000"
GREEN = "70AD47"
RED = "C00000"

# Fonts
FONT_TITLE = Font(name="Calibri", size=14, bold=True, color=WHITE)
FONT_HEADER = Font(name="Calibri", size=11, bold=True, color=WHITE)
FONT_SUBHEADER = Font(name="Calibri", size=10, bold=True, color=DARK_BLUE)
FONT_BODY = Font(name="Calibri", size=10, color=BLACK)
FONT_TOTAL = Font(name="Calibri", size=10, bold=True, color=BLACK)
FONT_KPI = Font(name="Calibri", size=11, bold=True, color=DARK_BLUE)

# Fills
FILL_NONE = PatternFill(fill_type=None)
FILL_TITLE = PatternFill(start_color=DARK_BLUE, end_color=DARK_BLUE, fill_type="solid")
FILL_HEADER = PatternFill(start_color=MEDIUM_BLUE, end_color=MEDIUM_BLUE, fill_type="solid")
FILL_SUBHEADER = PatternFill(start_color=LIGHT_BLUE, end_color=LIGHT_BLUE, fill_type="solid")
FILL_TOTAL = PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type="solid")
FILL_ALTERNATING = PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type="solid")

# Borders
THIN_BORDER = Side(style="thin", color=MEDIUM_BLUE)
MEDIUM_BORDER = Side(style="medium", color=DARK_BLUE)

BORDER_ALL = Border(
    left=THIN_BORDER, right=THIN_BORDER, top=THIN_BORDER, bottom=THIN_BORDER
)
BORDER_TOTAL = Border(
    left=THIN_BORDER, right=THIN_BORDER, top=MEDIUM_BORDER, bottom=MEDIUM_BORDER
)

# Alignments
ALIGN_LEFT = Alignment(horizontal="left", vertical="center")
ALIGN_CENTER = Alignment(horizontal="center", vertical="center")
ALIGN_RIGHT = Alignment(horizontal="right", vertical="center")

# Number formats
FORMAT_NUMBER = '#,##0'
FORMAT_NUMBER_K = '#,##0" k€"'
FORMAT_PERCENT = '0.0%'
FORMAT_PERCENT_VAR = '+0.0%;-0.0%;0.0%'
FORMAT_DAYS = '0" j"'

# Column widths
COL_WIDTH_LABEL = 35
COL_WIDTH_NUMBER = 15
COL_WIDTH_PERCENT = 12
