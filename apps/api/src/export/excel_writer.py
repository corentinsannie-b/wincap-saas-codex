"""Excel Databook writer with professional formatting."""

import json
from datetime import datetime
from decimal import Decimal
from pathlib import Path
from typing import Any, Dict, List, Optional, Union

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from src.models.financials import BalanceSheet, KPIs, ProfitLoss
from src.models.entry import JournalEntry

import sys
sys.path.insert(0, str(Path(__file__).parent.parent.parent))
from templates.excel.databook_style import (
    ALIGN_CENTER,
    ALIGN_LEFT,
    ALIGN_RIGHT,
    BORDER_ALL,
    BORDER_TOTAL,
    COL_WIDTH_LABEL,
    COL_WIDTH_NUMBER,
    COL_WIDTH_PERCENT,
    FILL_HEADER,
    FILL_NONE,
    FILL_SUBHEADER,
    FILL_TITLE,
    FILL_TOTAL,
    FONT_BODY,
    FONT_HEADER,
    FONT_SUBHEADER,
    FONT_TITLE,
    FONT_TOTAL,
    FORMAT_NUMBER,
    FORMAT_NUMBER_K,
    FORMAT_PERCENT,
    FORMAT_DAYS,
)


class ExcelWriter:
    """Generate Excel Databook from financial data."""

    def __init__(self, company_name: str = ""):
        self.company_name = company_name
        self.workbook = Workbook()
        # Remove default sheet
        self.workbook.remove(self.workbook.active)

    def generate(
        self,
        pl_list: List[ProfitLoss],
        balance_list: List[BalanceSheet],
        kpis_list: List[KPIs],
        output_path: Union[str, Path],
        entries: List[JournalEntry] = None,
        cashflows: List[dict] = None,
        monthly_data: dict = None,
        variance_data: dict = None,
        detail_data: dict = None,
    ) -> Path:
        """Generate complete Databook."""
        output_path = Path(output_path)

        # Create core sheets
        self._create_synthesis_sheet(kpis_list)
        self._create_key_figures_sheet(kpis_list)
        self._create_pl_sheet(pl_list)
        self._create_balance_sheet(balance_list)
        self._create_bfr_sheet(balance_list, kpis_list)
        self._create_kpis_sheet(kpis_list)

        # Create additional sheets if data provided
        if cashflows:
            self._create_cashflow_sheet(cashflows)

        if monthly_data:
            self._create_monthly_sheet(monthly_data)

        if variance_data and len(pl_list) >= 2:
            self._create_variance_sheet(variance_data, pl_list)
            self._create_ebitda_bridge_sheet(variance_data, pl_list)

        # Create detailed analysis sheets if data provided
        if detail_data:
            self._create_detail_sheets(detail_data)

        # Save
        self.workbook.save(output_path)
        return output_path

    def _variation_labels(self, years: List[int]) -> List[str]:
        """Build variation labels for adjacent years and multi-year delta."""
        if len(years) < 2:
            return []

        labels = [
            f"{str(curr)[-2:]}/{str(prev)[-2:]}"
            for curr, prev in zip(years[1:], years[:-1])
        ]
        if len(years) >= 3:
            labels.append(f"{str(years[-1])[-2:]}/{str(years[0])[-2:]}")

        return labels

    def _variation_pairs(self, years: List[int]) -> List[tuple]:
        """Return (current, previous) year pairs for variations."""
        if len(years) < 2:
            return []

        pairs = [(curr, prev) for curr, prev in zip(years[1:], years[:-1])]
        if len(years) >= 3:
            pairs.append((years[-1], years[0]))

        return pairs

    def _export_columns(self, years: List[int]) -> dict:
        """Compute column positions for export-style layout."""
        label_col = 2
        year_start = 3
        year_end = year_start + len(years) - 1
        spacer_col = year_end + 1
        variation_start = spacer_col + 1
        return {
            "label_col": label_col,
            "year_start": year_start,
            "year_end": year_end,
            "spacer_col": spacer_col,
            "variation_start": variation_start,
        }

    def _write_export_header(
        self,
        ws: Worksheet,
        start_row: int,
        years: List[int],
        unit_label: str = "En k€",
        include_variations: bool = True,
        year_labels: Optional[List[str]] = None,
    ) -> int:
        """Write export-style header block and return first data row."""
        cols = self._export_columns(years)
        var_labels = self._variation_labels(years) if include_variations else []
        header_year_labels = year_labels or [f"FY{year}" for year in years]

        header_row = start_row
        subheader_row = start_row + 1

        # Unit label
        cell = ws.cell(row=header_row, column=cols["label_col"], value=unit_label)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = ALIGN_CENTER
        cell.border = BORDER_ALL

        # Year headers
        for idx, label in enumerate(header_year_labels):
            col = cols["year_start"] + idx
            cell = ws.cell(row=header_row, column=col, value=label)
            cell.font = FONT_HEADER
            cell.fill = FILL_HEADER
            cell.alignment = ALIGN_CENTER
            cell.border = BORDER_ALL

            cell = ws.cell(row=subheader_row, column=col, value="Réel")
            cell.font = FONT_HEADER
            cell.fill = FILL_HEADER
            cell.alignment = ALIGN_CENTER
            cell.border = BORDER_ALL

        # Spacer column (kept blank but bordered for consistency)
        spacer = ws.cell(row=header_row, column=cols["spacer_col"], value="")
        spacer.border = BORDER_ALL
        spacer = ws.cell(row=subheader_row, column=cols["spacer_col"], value="")
        spacer.border = BORDER_ALL

        # Variations block
        if var_labels:
            cell = ws.cell(row=header_row, column=cols["variation_start"], value="Variations")
            cell.font = FONT_HEADER
            cell.fill = FILL_HEADER
            cell.alignment = ALIGN_CENTER
            cell.border = BORDER_ALL

            for idx, label in enumerate(var_labels):
                col = cols["variation_start"] + idx
                cell = ws.cell(row=subheader_row, column=col, value=label)
                cell.font = FONT_HEADER
                cell.fill = FILL_HEADER
                cell.alignment = ALIGN_CENTER
                cell.border = BORDER_ALL

        return start_row + 3  # header rows + blank row

    def _create_key_figures_sheet(self, kpis_list: List[KPIs]):
        """Create export-style key figures sheet."""
        ws = self.workbook.create_sheet("Chiffres clés")
        years = [kpi.year for kpi in kpis_list]
        cols = self._export_columns(years)

        # Title
        ws.merge_cells("A1:F1")
        ws["A1"] = "CHIFFRES CLÉS"
        ws["A1"].font = FONT_TITLE
        ws["A1"].fill = FILL_TITLE
        ws["A1"].alignment = ALIGN_CENTER

        data_row = self._write_export_header(ws, 5, years, unit_label="En k€", include_variations=True)

        indicators = [
            ("Chiffre d'affaires", "revenue", 1000, FORMAT_NUMBER),
            ("EBITDA", "ebitda", 1000, FORMAT_NUMBER),
            ("Marge EBITDA (%)", "ebitda_margin", 1, FORMAT_PERCENT),
            ("Résultat Net", "net_income", 1000, FORMAT_NUMBER),
            ("BFR", "working_capital", 1000, FORMAT_NUMBER),
            ("DSO (jours)", "dso", 1, FORMAT_DAYS),
            ("DPO (jours)", "dpo", 1, FORMAT_DAYS),
        ]

        var_pairs = self._variation_pairs(years)

        row = data_row
        for label, attr, divisor, fmt in indicators:
            cell = ws.cell(row=row, column=cols["label_col"], value=label)
            cell.font = FONT_BODY
            cell.alignment = ALIGN_LEFT
            cell.border = BORDER_ALL

            values_by_year = {}
            for idx, kpi in enumerate(kpis_list):
                value = getattr(kpi, attr, Decimal("0"))
                values_by_year[kpi.year] = value
                display_value = float(value / divisor) if divisor > 1 else float(value)
                if fmt == FORMAT_PERCENT:
                    display_value = float(value) / 100

                cell = ws.cell(row=row, column=cols["year_start"] + idx, value=display_value)
                cell.font = FONT_BODY
                cell.alignment = ALIGN_RIGHT
                cell.border = BORDER_ALL
                cell.number_format = fmt

            # Variations
            for idx, (curr, prev) in enumerate(var_pairs):
                prev_val = values_by_year.get(prev, Decimal("0"))
                curr_val = values_by_year.get(curr, Decimal("0"))
                var_pct = float((curr_val - prev_val) / abs(prev_val)) if prev_val != 0 else 0
                cell = ws.cell(row=row, column=cols["variation_start"] + idx, value=var_pct)
                cell.font = FONT_BODY
                cell.alignment = ALIGN_RIGHT
                cell.border = BORDER_ALL
                cell.number_format = FORMAT_PERCENT

            row += 1

        ws.column_dimensions["A"].width = 3
        ws.column_dimensions["B"].width = COL_WIDTH_LABEL
        for col in range(cols["year_start"], cols["variation_start"] + len(var_pairs) + 1):
            ws.column_dimensions[get_column_letter(col)].width = COL_WIDTH_NUMBER

    def _create_synthesis_sheet(self, kpis_list: List[KPIs]):
        """Create synthesis summary sheet."""
        ws = self.workbook.create_sheet("Synthèse")

        # Title
        ws.merge_cells("A1:F1")
        ws["A1"] = f"SYNTHÈSE FINANCIÈRE - {self.company_name}"
        ws["A1"].font = FONT_TITLE
        ws["A1"].fill = FILL_TITLE
        ws["A1"].alignment = ALIGN_CENTER

        # Headers
        headers = ["Indicateur"] + [f"FY{kpi.year}" for kpi in kpis_list]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = FONT_HEADER
            cell.fill = FILL_HEADER
            cell.alignment = ALIGN_CENTER
            cell.border = BORDER_ALL

        # Data rows
        indicators = [
            ("CA (k€)", "revenue", 1000, FORMAT_NUMBER),
            ("EBITDA (k€)", "ebitda", 1000, FORMAT_NUMBER),
            ("Marge EBITDA (%)", "ebitda_margin", 1, FORMAT_PERCENT),
            ("Résultat Net (k€)", "net_income", 1000, FORMAT_NUMBER),
            ("BFR (k€)", "working_capital", 1000, FORMAT_NUMBER),
            ("DSO (jours)", "dso", 1, FORMAT_DAYS),
            ("DPO (jours)", "dpo", 1, FORMAT_DAYS),
        ]

        for row_idx, (label, attr, divisor, fmt) in enumerate(indicators, 4):
            # Label
            cell = ws.cell(row=row_idx, column=1, value=label)
            cell.font = FONT_BODY
            cell.alignment = ALIGN_LEFT
            cell.border = BORDER_ALL

            # Values
            for col_idx, kpi in enumerate(kpis_list, 2):
                value = getattr(kpi, attr, Decimal("0"))
                if isinstance(value, Decimal):
                    if fmt == FORMAT_PERCENT:
                        value = float(value) / 100
                    elif divisor > 1:
                        value = float(value / divisor)
                    else:
                        value = float(value)

                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = FONT_BODY
                cell.alignment = ALIGN_RIGHT
                cell.border = BORDER_ALL
                cell.number_format = fmt

        # Column widths
        ws.column_dimensions["A"].width = COL_WIDTH_LABEL
        for col in range(2, len(kpis_list) + 2):
            ws.column_dimensions[get_column_letter(col)].width = COL_WIDTH_NUMBER

    def _create_pl_sheet(self, pl_list: List[ProfitLoss]):
        """Create P&L sheet."""
        ws = self.workbook.create_sheet("P&L")
        years = [pl.year for pl in pl_list]
        cols = self._export_columns(years)

        # Title
        ws.merge_cells("A1:F1")
        ws["A1"] = "COMPTE DE RÉSULTAT"
        ws["A1"].font = FONT_TITLE
        ws["A1"].fill = FILL_TITLE
        ws["A1"].alignment = ALIGN_CENTER

        data_row = self._write_export_header(ws, 5, years, unit_label="En k€", include_variations=True)

        # P&L lines
        pl_lines = [
            ("Chiffre d'affaires", "revenue", False),
            ("Autres produits", "other_revenue", False),
            ("Production", "production", True),
            ("", None, False),  # Empty row
            ("Achats", "purchases", False),
            ("Charges externes", "external_charges", False),
            ("Impôts et taxes", "taxes", False),
            ("Charges de personnel", "personnel", False),
            ("Autres charges", "other_charges", False),
            ("Total charges", "total_charges", True),
            ("", None, False),
            ("EBITDA", "ebitda", True),
            ("Marge EBITDA (%)", "ebitda_margin", True),
            ("", None, False),
            ("Dotations aux amortissements", "depreciation", False),
            ("EBIT", "ebit", True),
            ("", None, False),
            ("Résultat financier", "financial_result", False),
            ("Résultat exceptionnel", "exceptional_result", False),
            ("Impôt sur les sociétés", "income_tax", False),
            ("", None, False),
            ("Résultat Net", "net_income", True),
        ]

        row = data_row
        var_pairs = self._variation_pairs(years)
        for label, attr, is_total in pl_lines:
            if attr is None:
                row += 1
                continue

            # Label
            cell = ws.cell(row=row, column=cols["label_col"], value=label)
            cell.font = FONT_TOTAL if is_total else FONT_BODY
            if is_total:
                    cell.fill = FILL_TOTAL
            cell.alignment = ALIGN_LEFT
            cell.border = BORDER_TOTAL if is_total else BORDER_ALL

            # Values
            for col_idx, pl in enumerate(pl_list, 2):
                value = getattr(pl, attr, Decimal("0"))
                is_percent = "margin" in attr.lower() or "%" in label

                if is_percent:
                    display_value = float(value) / 100
                else:
                    display_value = float(value / 1000)

                cell = ws.cell(row=row, column=cols["year_start"] + col_idx - 2, value=display_value)
                cell.font = FONT_TOTAL if is_total else FONT_BODY
                if is_total:
                    cell.fill = FILL_TOTAL
                cell.alignment = ALIGN_RIGHT
                cell.border = BORDER_TOTAL if is_total else BORDER_ALL
                cell.number_format = FORMAT_PERCENT if is_percent else FORMAT_NUMBER

            # Variation column
            if var_pairs and not attr.endswith("margin"):
                values_by_year = {pl.year: getattr(pl, attr, Decimal("0")) for pl in pl_list}
                for idx, (curr, prev) in enumerate(var_pairs):
                    prev_val = values_by_year.get(prev, Decimal("0"))
                    curr_val = values_by_year.get(curr, Decimal("0"))
                    if prev_val != 0:
                        var_pct = float((curr_val - prev_val) / abs(prev_val))
                    else:
                        var_pct = 0
                    cell = ws.cell(row=row, column=cols["variation_start"] + idx, value=var_pct)
                    cell.font = FONT_BODY
                    cell.alignment = ALIGN_RIGHT
                    cell.border = BORDER_ALL
                    cell.number_format = FORMAT_PERCENT

            row += 1

        # Column widths
        ws.column_dimensions["A"].width = 3
        ws.column_dimensions["B"].width = COL_WIDTH_LABEL
        for col in range(cols["year_start"], cols["variation_start"] + len(var_pairs) + 1):
            ws.column_dimensions[get_column_letter(col)].width = COL_WIDTH_NUMBER

    def _create_balance_sheet(self, balance_list: List[BalanceSheet]):
        """Create Balance Sheet."""
        ws = self.workbook.create_sheet("Bilan")
        years = [b.year for b in balance_list]
        cols = self._export_columns(years)

        # Title
        ws.merge_cells("A1:F1")
        ws["A1"] = "BILAN"
        ws["A1"].font = FONT_TITLE
        ws["A1"].fill = FILL_TITLE
        ws["A1"].alignment = ALIGN_CENTER

        year_labels = [f"Dec-{str(year)[-2:]}" for year in years]
        data_row = self._write_export_header(
            ws,
            5,
            years,
            unit_label="En k€",
            include_variations=False,
            year_labels=year_labels,
        )

        # Balance lines
        balance_lines = [
            ("ACTIF", None, "section"),
            ("Immobilisations nettes", "fixed_assets", False),
            ("Stocks", "inventory", False),
            ("Créances clients", "receivables", False),
            ("Autres créances", "other_receivables", False),
            ("Trésorerie", "cash", False),
            ("Total Actif", "total_assets", True),
            ("", None, None),
            ("PASSIF", None, "section"),
            ("Capitaux propres", "equity", False),
            ("Provisions", "provisions", False),
            ("Dettes financières", "financial_debt", False),
            ("Dettes fournisseurs", "payables", False),
            ("Autres dettes", "other_payables", False),
            ("Total Passif", "total_liabilities", True),
        ]

        row = data_row
        for label, attr, style in balance_lines:
            if attr is None and style is None:
                row += 1
                continue

            # Label
            cell = ws.cell(row=row, column=cols["label_col"], value=label)
            if style == "section":
                cell.font = FONT_SUBHEADER
                cell.fill = FILL_SUBHEADER
            elif style:
                cell.font = FONT_TOTAL
                cell.fill = FILL_TOTAL
            else:
                cell.font = FONT_BODY
            cell.alignment = ALIGN_LEFT
            cell.border = BORDER_ALL

            # Values
            if attr:
                for col_idx, balance in enumerate(balance_list, 2):
                    value = getattr(balance, attr, Decimal("0"))
                    display_value = float(value / 1000)

                    cell = ws.cell(row=row, column=cols["year_start"] + col_idx - 2, value=display_value)
                    cell.font = FONT_TOTAL if style else FONT_BODY
                    if style:
                        cell.fill = FILL_TOTAL
                    cell.alignment = ALIGN_RIGHT
                    cell.border = BORDER_ALL
                    cell.number_format = FORMAT_NUMBER

            row += 1

        # Column widths
        ws.column_dimensions["A"].width = 3
        ws.column_dimensions["B"].width = COL_WIDTH_LABEL
        for col in range(cols["year_start"], cols["year_end"] + 1):
            ws.column_dimensions[get_column_letter(col)].width = COL_WIDTH_NUMBER

    def _create_bfr_sheet(self, balance_list: List[BalanceSheet], kpis_list: List[KPIs]):
        """Create BFR (Working Capital) analysis sheet."""
        ws = self.workbook.create_sheet("BFR")

        # Title
        ws.merge_cells("A1:F1")
        ws["A1"] = "ANALYSE DU BFR"
        ws["A1"].font = FONT_TITLE
        ws["A1"].fill = FILL_TITLE
        ws["A1"].alignment = ALIGN_CENTER

        # Headers
        headers = ["Libellé"] + [f"Dec-{str(b.year)[-2:]}" for b in balance_list]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = FONT_HEADER
            cell.fill = FILL_HEADER
            cell.alignment = ALIGN_CENTER
            cell.border = BORDER_ALL

        # BFR components
        bfr_lines = [
            ("Stocks (k€)", "inventory", 1000, FORMAT_NUMBER),
            ("Créances clients (k€)", "receivables", 1000, FORMAT_NUMBER),
            ("Dettes fournisseurs (k€)", "payables", 1000, FORMAT_NUMBER),
            ("BFR (k€)", "working_capital", 1000, FORMAT_NUMBER),
        ]

        row = 4
        for label, attr, divisor, fmt in bfr_lines:
            is_bfr_total = "BFR" in label
            cell = ws.cell(row=row, column=1, value=label)
            cell.font = FONT_TOTAL if is_bfr_total else FONT_BODY
            if is_bfr_total:
                cell.fill = FILL_TOTAL
            cell.alignment = ALIGN_LEFT
            cell.border = BORDER_ALL

            for col_idx, balance in enumerate(balance_list, 2):
                value = getattr(balance, attr, Decimal("0"))
                display_value = float(value / divisor)

                cell = ws.cell(row=row, column=col_idx, value=display_value)
                cell.font = FONT_TOTAL if is_bfr_total else FONT_BODY
                if is_bfr_total:
                    cell.fill = FILL_TOTAL
                cell.alignment = ALIGN_RIGHT
                cell.border = BORDER_ALL
                cell.number_format = fmt

            row += 1

        # DSO/DPO section
        row += 1
        kpis_by_year = {kpi.year: kpi for kpi in kpis_list}

        dso_dpo_lines = [
            ("DSO (jours)", "dso", 1, FORMAT_DAYS),
            ("DPO (jours)", "dpo", 1, FORMAT_DAYS),
        ]

        for label, attr, divisor, fmt in dso_dpo_lines:
            cell = ws.cell(row=row, column=1, value=label)
            cell.font = FONT_BODY
            cell.alignment = ALIGN_LEFT
            cell.border = BORDER_ALL

            for col_idx, balance in enumerate(balance_list, 2):
                kpi = kpis_by_year.get(balance.year)
                if kpi:
                    value = getattr(kpi, attr, Decimal("0"))
                    display_value = float(value)

                    cell = ws.cell(row=row, column=col_idx, value=display_value)
                    cell.font = FONT_BODY
                    cell.alignment = ALIGN_RIGHT
                    cell.border = BORDER_ALL
                    cell.number_format = fmt

            row += 1

        # Column widths
        ws.column_dimensions["A"].width = COL_WIDTH_LABEL
        for col in range(2, len(balance_list) + 2):
            ws.column_dimensions[get_column_letter(col)].width = COL_WIDTH_NUMBER

    def _create_kpis_sheet(self, kpis_list: List[KPIs]):
        """Create KPIs sheet."""
        ws = self.workbook.create_sheet("KPIs")

        # Title
        ws.merge_cells("A1:F1")
        ws["A1"] = "INDICATEURS CLÉS DE PERFORMANCE"
        ws["A1"].font = FONT_TITLE
        ws["A1"].fill = FILL_TITLE
        ws["A1"].alignment = ALIGN_CENTER

        # Headers
        headers = ["Indicateur"] + [f"FY{kpi.year}" for kpi in kpis_list]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = FONT_HEADER
            cell.fill = FILL_HEADER
            cell.alignment = ALIGN_CENTER
            cell.border = BORDER_ALL

        # KPI rows
        kpi_lines = [
            ("Rentabilité", None, "section"),
            ("CA (k€)", "revenue", 1000, FORMAT_NUMBER),
            ("EBITDA (k€)", "ebitda", 1000, FORMAT_NUMBER),
            ("Marge EBITDA", "ebitda_margin", 100, FORMAT_PERCENT),
            ("Résultat Net (k€)", "net_income", 1000, FORMAT_NUMBER),
            ("", None, None, None),
            ("Trésorerie", None, "section"),
            ("BFR (k€)", "working_capital", 1000, FORMAT_NUMBER),
            ("Dette nette (k€)", "net_debt", 1000, FORMAT_NUMBER),
            ("", None, None, None),
            ("Rotation", None, "section"),
            ("DSO (jours)", "dso", 1, FORMAT_DAYS),
            ("DPO (jours)", "dpo", 1, FORMAT_DAYS),
            ("DIO (jours)", "dio", 1, FORMAT_DAYS),
        ]

        row = 4
        for item in kpi_lines:
            if len(item) == 3:  # Section header
                label, _, style = item
                cell = ws.cell(row=row, column=1, value=label)
                cell.font = FONT_SUBHEADER
                cell.fill = FILL_SUBHEADER
                cell.alignment = ALIGN_LEFT
                cell.border = BORDER_ALL
                row += 1
                continue

            label, attr, divisor, fmt = item
            if attr is None:
                row += 1
                continue

            cell = ws.cell(row=row, column=1, value=label)
            cell.font = FONT_BODY
            cell.alignment = ALIGN_LEFT
            cell.border = BORDER_ALL

            for col_idx, kpi in enumerate(kpis_list, 2):
                value = getattr(kpi, attr, Decimal("0"))
                if divisor == 100:  # Percentage
                    display_value = float(value) / 100
                elif divisor > 1:
                    display_value = float(value / divisor)
                else:
                    display_value = float(value)

                cell = ws.cell(row=row, column=col_idx, value=display_value)
                cell.font = FONT_BODY
                cell.alignment = ALIGN_RIGHT
                cell.border = BORDER_ALL
                cell.number_format = fmt

            row += 1

        # Column widths
        ws.column_dimensions["A"].width = COL_WIDTH_LABEL
        for col in range(2, len(kpis_list) + 2):
            ws.column_dimensions[get_column_letter(col)].width = COL_WIDTH_NUMBER

    def _create_cashflow_sheet(self, cashflows: List[dict]):
        """Create Cash-Flow statement sheet."""
        ws = self.workbook.create_sheet("Cash-Flow")
        years = [cf["year"] for cf in cashflows]
        cols = self._export_columns(years)

        # Title
        ws.merge_cells("A1:F1")
        ws["A1"] = "TABLEAU DES FLUX DE TRÉSORERIE"
        ws["A1"].font = FONT_TITLE
        ws["A1"].fill = FILL_TITLE
        ws["A1"].alignment = ALIGN_CENTER

        data_row = self._write_export_header(ws, 5, years, unit_label="En k€", include_variations=False)
        cumulative_col = cols["year_end"] + 2
        cumulative_label = ""
        if years:
            cumulative_label = f"FY{years[0]}-{years[-1]}"

        cell = ws.cell(row=5, column=cumulative_col, value=cumulative_label)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = ALIGN_CENTER
        cell.border = BORDER_ALL

        cell = ws.cell(row=6, column=cumulative_col, value="24m.")
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = ALIGN_CENTER
        cell.border = BORDER_ALL

        # Cash-Flow lines
        cf_lines = [
            ("FLUX D'EXPLOITATION", None, "section"),
            ("EBITDA", "ebitda", False),
            ("Variation créances clients", "var_receivables", False),
            ("Variation stocks", "var_inventory", False),
            ("Variation dettes fournisseurs", "var_payables", False),
            ("Variation autres éléments BFR", "var_other_wc", False),
            ("Variation du BFR", "var_bfr", True),
            ("Flux de trésorerie d'exploitation", "operating_cf", True),
            ("", None, None),
            ("FLUX D'INVESTISSEMENT", None, "section"),
            ("Investissements (CAPEX)", "capex", False),
            ("Flux de trésorerie d'investissement", "investing_cf", True),
            ("", None, None),
            ("FLUX DE FINANCEMENT", None, "section"),
            ("Variation dette financière", "var_debt", False),
            ("Variation capitaux propres", "var_equity", False),
            ("Flux de trésorerie de financement", "financing_cf", True),
            ("", None, None),
            ("VARIATION DE TRÉSORERIE", "net_cash_change", True),
            ("Trésorerie début de période", "cash_start", False),
            ("Trésorerie fin de période", "cash_end", True),
        ]

        row = data_row
        for label, key, style in cf_lines:
            if key is None and style is None:
                row += 1
                continue

            cell = ws.cell(row=row, column=cols["label_col"], value=label)
            if style == "section":
                cell.font = FONT_SUBHEADER
                cell.fill = FILL_SUBHEADER
            elif style:
                cell.font = FONT_TOTAL
                cell.fill = FILL_TOTAL
            else:
                cell.font = FONT_BODY
            cell.alignment = ALIGN_LEFT
            cell.border = BORDER_ALL

            if key:
                total_value = Decimal("0")
                for col_idx, cf in enumerate(cashflows, 2):
                    value = cf.get(key, Decimal("0"))
                    display_value = float(value / 1000) if isinstance(value, Decimal) else float(value) / 1000
                    if isinstance(value, Decimal):
                        total_value += value

                    cell = ws.cell(row=row, column=cols["year_start"] + col_idx - 2, value=display_value)
                    cell.font = FONT_TOTAL if style else FONT_BODY
                    if style and style != "section":
                        cell.fill = FILL_TOTAL
                    cell.alignment = ALIGN_RIGHT
                    cell.border = BORDER_ALL
                    cell.number_format = FORMAT_NUMBER

                cell = ws.cell(row=row, column=cumulative_col, value=float(total_value / 1000))
                cell.font = FONT_TOTAL if style else FONT_BODY
                if style and style != "section":
                    cell.fill = FILL_TOTAL
                cell.alignment = ALIGN_RIGHT
                cell.border = BORDER_ALL
                cell.number_format = FORMAT_NUMBER

            row += 1

        # Column widths
        ws.column_dimensions["A"].width = 3
        ws.column_dimensions["B"].width = COL_WIDTH_LABEL + 5
        for col in range(cols["year_start"], cumulative_col + 1):
            ws.column_dimensions[get_column_letter(col)].width = COL_WIDTH_NUMBER

    def _create_monthly_sheet(self, monthly_data: dict):
        """Create monthly analysis sheet."""
        ws = self.workbook.create_sheet("Mensuel")

        # Title
        ws.merge_cells("A1:N1")
        ws["A1"] = "ANALYSE MENSUELLE DU CHIFFRE D'AFFAIRES"
        ws["A1"].font = FONT_TITLE
        ws["A1"].fill = FILL_TITLE
        ws["A1"].alignment = ALIGN_CENTER

        # Month names
        month_names = ["Jan", "Fév", "Mar", "Avr", "Mai", "Juin",
                       "Juil", "Août", "Sep", "Oct", "Nov", "Déc", "Total"]

        # Headers
        headers = ["Année"] + month_names
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = FONT_HEADER
            cell.fill = FILL_HEADER
            cell.alignment = ALIGN_CENTER
            cell.border = BORDER_ALL

        # Data rows
        revenue_data = monthly_data.get("revenue", {})
        row = 4

        for year in sorted(revenue_data.keys()):
            months = revenue_data[year]

            # Year label
            cell = ws.cell(row=row, column=1, value=f"FY{year}")
            cell.font = FONT_BODY
            cell.alignment = ALIGN_LEFT
            cell.border = BORDER_ALL

            # Monthly values
            year_total = Decimal("0")
            for month in range(1, 13):
                value = months.get(month, Decimal("0"))
                year_total += value
                display_value = float(value / 1000)

                cell = ws.cell(row=row, column=month + 1, value=display_value)
                cell.font = FONT_BODY
                cell.alignment = ALIGN_RIGHT
                cell.border = BORDER_ALL
                cell.number_format = FORMAT_NUMBER

            # Year total
            cell = ws.cell(row=row, column=14, value=float(year_total / 1000))
            cell.font = FONT_TOTAL
            cell.fill = FILL_TOTAL
            cell.alignment = ALIGN_RIGHT
            cell.border = BORDER_ALL
            cell.number_format = FORMAT_NUMBER

            row += 1

        # Column widths
        ws.column_dimensions["A"].width = 12
        for col in range(2, 15):
            ws.column_dimensions[get_column_letter(col)].width = 10

    def _create_variance_sheet(self, variance_data: dict, pl_list: List[ProfitLoss]):
        """Create variance analysis sheet."""
        ws = self.workbook.create_sheet("Variances")

        prev_year = pl_list[-2].year
        curr_year = pl_list[-1].year

        # Title
        ws.merge_cells("A1:F1")
        ws["A1"] = f"ANALYSE DES ÉCARTS FY{prev_year} → FY{curr_year}"
        ws["A1"].font = FONT_TITLE
        ws["A1"].fill = FILL_TITLE
        ws["A1"].alignment = ALIGN_CENTER

        # Headers
        headers = ["Poste", f"FY{prev_year} (k€)", f"FY{curr_year} (k€)", "Écart (k€)", "Écart (%)"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = FONT_HEADER
            cell.fill = FILL_HEADER
            cell.alignment = ALIGN_CENTER
            cell.border = BORDER_ALL

        # Variance lines from data
        cost_breakdown = variance_data.get("cost_breakdown", [])
        row = 4

        for item in cost_breakdown:
            # Label
            cell = ws.cell(row=row, column=1, value=item["label"])
            cell.font = FONT_BODY
            cell.alignment = ALIGN_LEFT
            cell.border = BORDER_ALL

            # Previous year
            cell = ws.cell(row=row, column=2, value=item["prev_value"])
            cell.font = FONT_BODY
            cell.alignment = ALIGN_RIGHT
            cell.border = BORDER_ALL
            cell.number_format = FORMAT_NUMBER

            # Current year
            cell = ws.cell(row=row, column=3, value=item["curr_value"])
            cell.font = FONT_BODY
            cell.alignment = ALIGN_RIGHT
            cell.border = BORDER_ALL
            cell.number_format = FORMAT_NUMBER

            # Absolute variance
            cell = ws.cell(row=row, column=4, value=item["var_abs"])
            cell.font = FONT_BODY
            cell.alignment = ALIGN_RIGHT
            cell.border = BORDER_ALL
            cell.number_format = FORMAT_NUMBER

            # Percentage variance
            cell = ws.cell(row=row, column=5, value=item["var_pct"] / 100 if item["var_pct"] else 0)
            cell.font = FONT_BODY
            cell.alignment = ALIGN_RIGHT
            cell.border = BORDER_ALL
            cell.number_format = FORMAT_PERCENT

            row += 1

        # Add EBITDA bridge section
        row += 2
        cell = ws.cell(row=row, column=1, value="PONT EBITDA")
        cell.font = FONT_SUBHEADER
        cell.fill = FILL_SUBHEADER
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        row += 1

        bridge = variance_data.get("ebitda_bridge", [])
        for item in bridge:
            cell = ws.cell(row=row, column=1, value=item["label"])
            cell.font = FONT_TOTAL if item["type"] in ("start", "end") else FONT_BODY
            if item["type"] in ("start", "end"):
                cell.fill = FILL_TOTAL
            cell.alignment = ALIGN_LEFT
            cell.border = BORDER_ALL

            cell = ws.cell(row=row, column=2, value=item["value"])
            cell.font = FONT_TOTAL if item["type"] in ("start", "end") else FONT_BODY
            if item["type"] in ("start", "end"):
                cell.fill = FILL_TOTAL
            cell.alignment = ALIGN_RIGHT
            cell.border = BORDER_ALL
            cell.number_format = FORMAT_NUMBER

            row += 1

        # Column widths
        ws.column_dimensions["A"].width = COL_WIDTH_LABEL
        for col in range(2, 6):
            ws.column_dimensions[get_column_letter(col)].width = COL_WIDTH_NUMBER

    def _create_ebitda_bridge_sheet(self, variance_data: dict, pl_list: List[ProfitLoss]):
        """Create EBITDA bridge sheet to match export structure."""
        ws = self.workbook.create_sheet("Bridge EBITDA")

        # Title
        ws.merge_cells("A1:F1")
        ws["A1"] = "BRIDGE EBITDA"
        ws["A1"].font = FONT_TITLE
        ws["A1"].fill = FILL_TITLE
        ws["A1"].alignment = ALIGN_CENTER

        # Header row (export-style)
        header_row = 5
        cell = ws.cell(row=header_row, column=2, value="En k€")
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = ALIGN_CENTER
        cell.border = BORDER_ALL

        cell = ws.cell(row=header_row, column=3, value="Variations")
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = ALIGN_CENTER
        cell.border = BORDER_ALL

        cell = ws.cell(row=header_row, column=5, value="MSCD")
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = ALIGN_CENTER
        cell.border = BORDER_ALL

        cell = ws.cell(row=header_row, column=6, value="Var. 2y")
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = ALIGN_CENTER
        cell.border = BORDER_ALL

        bridge = variance_data.get("ebitda_bridge", [])
        row = header_row + 2
        for item in bridge:
            cell = ws.cell(row=row, column=2, value=item.get("label"))
            cell.font = FONT_TOTAL if item.get("type") in ("start", "end") else FONT_BODY
            cell.alignment = ALIGN_LEFT
            cell.border = BORDER_ALL

            cell = ws.cell(row=row, column=3, value=item.get("value"))
            cell.font = FONT_TOTAL if item.get("type") in ("start", "end") else FONT_BODY
            cell.alignment = ALIGN_RIGHT
            cell.border = BORDER_ALL
            cell.number_format = FORMAT_NUMBER

            row += 1

        ws.column_dimensions["A"].width = 3
        ws.column_dimensions["B"].width = COL_WIDTH_LABEL
        for col in range(3, 7):
            ws.column_dimensions[get_column_letter(col)].width = COL_WIDTH_NUMBER

    def _create_detail_sheets(self, detail_data: dict):
        """Create detailed analysis sheets when --detailed flag is used."""
        # Top Accounts sheet
        if "top_accounts" in detail_data and detail_data["top_accounts"]:
            ws = self.workbook.create_sheet("Top Comptes")

            # Title
            ws.merge_cells("A1:E1")
            ws["A1"] = "TOP 20 COMPTES PAR VOLUME"
            ws["A1"].font = FONT_TITLE
            ws["A1"].fill = FILL_TITLE
            ws["A1"].alignment = ALIGN_CENTER

            # Headers
            headers = ["Compte", "Libellé", "Débit (k€)", "Crédit (k€)", "Volume (k€)"]
            for col, header in enumerate(headers, start=1):
                cell = ws.cell(row=3, column=col, value=header)
                cell.font = FONT_HEADER
                cell.fill = FILL_HEADER
                cell.alignment = ALIGN_CENTER
                cell.border = BORDER_ALL

            # Data rows
            row = 4
            for account in detail_data["top_accounts"]:
                ws.cell(row=row, column=1, value=account.get("account_num", "")).border = BORDER_ALL
                ws.cell(row=row, column=2, value=account.get("label", "")).border = BORDER_ALL

                cell = ws.cell(row=row, column=3, value=float(account.get("total_debit", 0)) / 1000)
                cell.number_format = FORMAT_NUMBER
                cell.alignment = ALIGN_RIGHT
                cell.border = BORDER_ALL

                cell = ws.cell(row=row, column=4, value=float(account.get("total_credit", 0)) / 1000)
                cell.number_format = FORMAT_NUMBER
                cell.alignment = ALIGN_RIGHT
                cell.border = BORDER_ALL

                cell = ws.cell(row=row, column=5, value=float(account.get("volume", 0)) / 1000)
                cell.number_format = FORMAT_NUMBER
                cell.alignment = ALIGN_RIGHT
                cell.border = BORDER_ALL

                row += 1

            ws.column_dimensions["A"].width = 12
            ws.column_dimensions["B"].width = 40
            for col in range(3, 6):
                ws.column_dimensions[get_column_letter(col)].width = COL_WIDTH_NUMBER

        # Category Breakdown sheet
        if "category_breakdown" in detail_data and detail_data["category_breakdown"]:
            ws = self.workbook.create_sheet("Détail Catégories")

            ws.merge_cells("A1:D1")
            ws["A1"] = "VENTILATION PAR CATÉGORIE"
            ws["A1"].font = FONT_TITLE
            ws["A1"].fill = FILL_TITLE
            ws["A1"].alignment = ALIGN_CENTER

            headers = ["Catégorie", "Débit (k€)", "Crédit (k€)", "Solde (k€)"]
            for col, header in enumerate(headers, start=1):
                cell = ws.cell(row=3, column=col, value=header)
                cell.font = FONT_HEADER
                cell.fill = FILL_HEADER
                cell.alignment = ALIGN_CENTER
                cell.border = BORDER_ALL

            row = 4
            for category, data in detail_data["category_breakdown"].items():
                ws.cell(row=row, column=1, value=category).border = BORDER_ALL

                cell = ws.cell(row=row, column=2, value=float(data.get("debit", 0)) / 1000)
                cell.number_format = FORMAT_NUMBER
                cell.alignment = ALIGN_RIGHT
                cell.border = BORDER_ALL

                cell = ws.cell(row=row, column=3, value=float(data.get("credit", 0)) / 1000)
                cell.number_format = FORMAT_NUMBER
                cell.alignment = ALIGN_RIGHT
                cell.border = BORDER_ALL

                cell = ws.cell(row=row, column=4, value=float(data.get("balance", 0)) / 1000)
                cell.number_format = FORMAT_NUMBER
                cell.alignment = ALIGN_RIGHT
                cell.border = BORDER_ALL

                row += 1

            ws.column_dimensions["A"].width = 30
            for col in range(2, 5):
                ws.column_dimensions[get_column_letter(col)].width = COL_WIDTH_NUMBER


def export_dashboard_json(
    pl_list: List[ProfitLoss],
    balance_list: List[BalanceSheet],
    kpis_list: List[KPIs],
    company_name: str,
    output_path: Union[str, Path],
    monthly_data: Optional[Dict] = None,
    variance_data: Optional[Dict] = None,
) -> Path:
    """Export financial data as JSON for dashboard consumption.

    Args:
        pl_list: List of ProfitLoss for each year
        balance_list: List of BalanceSheet for each year
        kpis_list: List of KPIs for each year
        company_name: Company name
        output_path: Path to save JSON file
        monthly_data: Optional monthly breakdown data
        variance_data: Optional variance analysis data

    Returns:
        Path to the generated JSON file
    """
    output_path = Path(output_path)

    def decimal_to_float(obj: Any) -> Any:
        """Convert Decimal to float for JSON serialization."""
        if isinstance(obj, Decimal):
            return float(obj)
        elif isinstance(obj, dict):
            return {k: decimal_to_float(v) for k, v in obj.items()}
        elif isinstance(obj, list):
            return [decimal_to_float(item) for item in obj]
        return obj

    # Build dashboard data structure
    dashboard_data = {
        "metadata": {
            "company_name": company_name,
            "generated_at": datetime.now().isoformat(),
            "years": [pl.year for pl in pl_list],
        },
        "pl": [
            {
                "year": pl.year,
                "revenue": float(pl.revenue),
                "production": float(pl.production),
                "purchases": float(pl.purchases),
                "external_charges": float(pl.external_charges),
                "personnel": float(pl.personnel),
                "taxes": float(pl.taxes),
                "other_charges": float(pl.other_charges),
                "ebitda": float(pl.ebitda),
                "ebitda_margin": float(pl.ebitda_margin),
                "depreciation": float(pl.depreciation),
                "ebit": float(pl.ebit),
                "financial_result": float(pl.financial_result),
                "exceptional_result": float(pl.exceptional_result),
                "income_tax": float(pl.income_tax),
                "net_income": float(pl.net_income),
            }
            for pl in pl_list
        ],
        "balance": [
            {
                "year": b.year,
                "fixed_assets": float(b.fixed_assets),
                "inventory": float(b.inventory),
                "receivables": float(b.receivables),
                "other_receivables": float(b.other_receivables),
                "cash": float(b.cash),
                "total_assets": float(b.total_assets),
                "equity": float(b.equity),
                "provisions": float(b.provisions),
                "financial_debt": float(b.financial_debt),
                "payables": float(b.payables),
                "other_payables": float(b.other_payables),
                "total_liabilities": float(b.total_liabilities),
                "working_capital": float(b.working_capital),
                "net_debt": float(b.net_debt),
            }
            for b in balance_list
        ],
        "kpis": [
            {
                "year": k.year,
                "revenue": float(k.revenue),
                "ebitda": float(k.ebitda),
                "ebitda_margin": float(k.ebitda_margin),
                "net_income": float(k.net_income),
                "working_capital": float(k.working_capital),
                "net_debt": float(k.net_debt),
                "dso": float(k.dso),
                "dpo": float(k.dpo),
                "dio": float(k.dio),
                "adjusted_ebitda": float(k.adjusted_ebitda),
            }
            for k in kpis_list
        ],
    }

    # Add monthly data if available
    if monthly_data:
        dashboard_data["monthly"] = decimal_to_float(monthly_data)

    # Add variance data if available
    if variance_data:
        dashboard_data["variance"] = decimal_to_float(variance_data)

    # Write JSON file
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(dashboard_data, f, indent=2, ensure_ascii=False)

    return output_path
