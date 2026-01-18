"""Template-based Excel writer that uses client Databook template."""

import shutil
from datetime import datetime
from decimal import Decimal
from pathlib import Path
from typing import List, Optional, Union

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from src.models.entry import JournalEntry


class TemplateWriter:
    """Generate Databook by populating a template with FEC data.

    This writer takes an existing client template and populates the FEC sheet
    with new data. All other sheets use formulas that reference FEC data,
    so they automatically update.
    """

    # FEC sheet structure
    FEC_HEADER_ROW = 9
    FEC_DATA_START_ROW = 11

    # Data columns to populate (non-formula columns)
    FEC_DATA_COLUMNS = {
        "JournalCode": 2,
        "JournalLib": 3,
        "EcritureNum": 4,
        "EcritureDate": 5,
        "Exercice": 6,
        "Date": 7,
        "Mois": 8,
        "CompteNum": 9,
        # Columns 10-14 are formula columns (XLOOKUP)
        "CompteLib": 15,
        "CompAuxNum": 16,
        "CompAuxLib": 17,
        "PieceRef": 18,
        "PieceDate": 19,
        "EcritureLib": 20,
        "Debit": 21,
        "Credit": 22,
        # Column 23 (Solde) is formula
        # Column 24 (P&L/Bilan) is formula
        "EcritureLet": 25,
        "DateLet": 26,
        "ValidDate": 27,
    }

    # Formula templates for calculated columns
    FEC_FORMULAS = {
        7: '=IFERROR(\n    IF($C{row}="A Nouveaux Détaillés",\n        DATE(2000 + RIGHT(F{row}, 2), 1, 1),\n        DATE(LEFT(E{row}, 4), MID(E{row}, 5, 2), RIGHT(E{row}, 2))\n    ),\n    ""\n)',  # Date
        8: '=EOMONTH(G{row},0)',  # Mois
        10: '=_xlfn.XLOOKUP($I{row},BG!$B:$B,BG!L:L)',  # Mapping 1
        11: '=_xlfn.XLOOKUP($I{row},BG!$B:$B,BG!M:M)',  # Mapping 2
        12: '=_xlfn.XLOOKUP($I{row},BG!$B:$B,BG!N:N)',  # Mapping 3
        13: '=_xlfn.XLOOKUP($I{row},BG!$B:$B,BG!O:O)',  # Mapping CF
        23: '=IF(LEFT(I{row},1) >= "6", V{row} - U{row}, U{row} - V{row})/1000',  # Solde
        24: '=IF(LEFT(I{row},1) >= "6", "P&L", "Bilan")',  # P&L/Bilan
    }

    def __init__(self, template_path: Union[str, Path]):
        """Initialize with path to template file.

        Args:
            template_path: Path to the client's template Excel file
        """
        self.template_path = Path(template_path)
        if not self.template_path.exists():
            raise FileNotFoundError(f"Template not found: {template_path}")

    def generate(
        self,
        entries: List[JournalEntry],
        output_path: Union[str, Path],
        company_name: str = "",
        years: Optional[List[int]] = None,
    ) -> Path:
        """Generate Databook by populating template with FEC entries.

        Args:
            entries: List of parsed FEC journal entries
            output_path: Where to save the generated file
            company_name: Company name for Cover sheet
            years: Optional list of years to filter entries

        Returns:
            Path to the generated file
        """
        output_path = Path(output_path)

        # Copy template to output location
        shutil.copy2(self.template_path, output_path)

        # Open the copied workbook
        wb = load_workbook(output_path)

        # Update Cover sheet if it exists
        if "Cover" in wb.sheetnames and company_name:
            self._update_cover(wb["Cover"], company_name)

        # Filter entries by years if specified
        if years:
            entries = [e for e in entries if e.effective_year in years]

        # Populate FEC sheet
        if "FEC" in wb.sheetnames:
            self._populate_fec(wb["FEC"], entries)

        # Save the workbook
        wb.save(output_path)

        return output_path

    def _update_cover(self, ws: Worksheet, company_name: str):
        """Update Cover sheet with company name."""
        # Company name is typically in cell B19 based on template analysis
        ws["B19"] = company_name

    def _populate_fec(self, ws: Worksheet, entries: List[JournalEntry]):
        """Populate FEC sheet with journal entries.

        Clears existing data rows and writes new entries with formulas.
        """
        # Clear existing data rows (keep headers and first formula row as reference)
        max_row = ws.max_row
        if max_row > self.FEC_DATA_START_ROW:
            # Delete all data rows
            ws.delete_rows(self.FEC_DATA_START_ROW, max_row - self.FEC_DATA_START_ROW + 1)

        # Sort entries by date and account
        sorted_entries = sorted(entries, key=lambda e: (e.effective_year, e.date, e.account_num))

        # Write new data
        row = self.FEC_DATA_START_ROW
        for entry in sorted_entries:
            self._write_fec_row(ws, row, entry)
            row += 1

        print(f"  Wrote {len(sorted_entries)} FEC entries to template")

    def _write_fec_row(self, ws: Worksheet, row: int, entry: JournalEntry):
        """Write a single FEC entry to a row with data and formulas."""
        # Data columns
        ws.cell(row=row, column=2, value=self._get_journal_code(entry))
        ws.cell(row=row, column=3, value=self._get_journal_lib(entry))
        ws.cell(row=row, column=4, value=row - self.FEC_DATA_START_ROW + 1)  # EcritureNum
        ws.cell(row=row, column=5, value=int(entry.date.strftime("%Y%m%d")))  # EcritureDate
        ws.cell(row=row, column=6, value=f"FY{str(entry.effective_year)[-2:]}")  # Exercice
        # Columns 7-8 (Date, Mois) are now formulas set below
        # CompteNum - keep as string to match BG sheet text values for XLOOKUP
        ws.cell(row=row, column=9, value=entry.account_num)

        # Formula columns (10-14, 23, 24)
        for col, formula_template in self.FEC_FORMULAS.items():
            ws.cell(row=row, column=col, value=formula_template.format(row=row))

        # More data columns
        ws.cell(row=row, column=15, value=entry.label)  # CompteLib
        ws.cell(row=row, column=16, value="")  # CompAuxNum
        ws.cell(row=row, column=17, value="")  # CompAuxLib
        ws.cell(row=row, column=18, value="")  # PieceRef
        ws.cell(row=row, column=19, value=int(entry.date.strftime("%Y%m%d")))  # PieceDate
        ws.cell(row=row, column=20, value=entry.label)  # EcritureLib
        ws.cell(row=row, column=21, value=float(entry.debit))  # Debit
        ws.cell(row=row, column=22, value=float(entry.credit))  # Credit
        # Column 23 and 24 are formulas (already set above)
        ws.cell(row=row, column=25, value="")  # EcritureLet
        ws.cell(row=row, column=26, value="")  # DateLet
        ws.cell(row=row, column=27, value=int(entry.date.strftime("%Y%m%d")))  # ValidDate

    def _get_journal_code(self, entry: JournalEntry) -> str:
        """Determine journal code based on account class."""
        account_class = entry.account_num[0] if entry.account_num else ""
        if account_class == "6":
            return "AC"  # Achats/Charges
        elif account_class == "7":
            return "VE"  # Ventes
        else:
            return "OD"  # Opérations Diverses

    def _get_journal_lib(self, entry: JournalEntry) -> str:
        """Determine journal label based on account class."""
        account_class = entry.account_num[0] if entry.account_num else ""
        if account_class == "6":
            return "Achats"
        elif account_class == "7":
            return "Ventes"
        else:
            return "Opérations Diverses"
